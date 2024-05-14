from cvzone.FaceDetectionModule import FaceDetector
import os
import pickle
import bbox
import numpy as np
import cvzone
import cv2
import face_recognition
import datetime
from datetime import datetime
import firebase_admin
from firebase_admin import credentials
from firebase_admin import db
from firebase_admin import storage
import xlsxwriter
from xlrd import open_workbook
import csv
import pandas as pd
import openpyxl
from openpyxl import Workbook
from datetime import datetime
import emailsending
import schedule
import time
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
import pandas as pd

cred = credentials.Certificate("serviceAccountKey.json")
firebase_admin.initialize_app(cred, {
    'databaseURL': 'https://final-face-recog-attendance-default-rtdb.firebaseio.com/',
    'storageBucket': 'final-face-recog-attendance.appspot.com'
})

bucket = storage.bucket()
cap = cv2.VideoCapture(0)

cap.set(3, 640)
cap.set(4, 480)

camUI = cv2.imread('resources/background.png')

# Importing the files images into a list
folderFilesPath = 'resources/FILES'
filesPathList = os.listdir(folderFilesPath)
imgFilesList = []
for path in filesPathList:
    imgFilesList.append(cv2.imread(os.path.join(folderFilesPath, path)))

# encoding file
print("loading")
file = open('DneEncodeFile.pickle', 'rb')
encodeListSearchedWithIds = pickle.load(file)
file.close()
encodeListSearched, studentsId = encodeListSearchedWithIds

# print(studentsId)
print("saved encoded files!")

imgFileType = 0
counter = 0
id = -1
imageStudents = []

# Create an Excel workbook and set up the worksheet
wb = openpyxl.Workbook()
ws = wb.active
ws.append(["ID", "Name", "Total Attendance", "Last Attendance Time"])

# Dictionary to store the last attendance time for each student
last_attendance_time_dict = {}


while True:

    success, img = cap.read()

    imgSize = cv2.resize(img, (0, 0), None, 0.25, 0.25)
    imgSize = cv2.cvtColor(imgSize, cv2.COLOR_BGR2RGB)
    # current frame
    facesFrame = face_recognition.face_locations(imgSize)
    eFrame = face_recognition.face_encodings(imgSize, facesFrame)

    camUI[162:162 + 480, 55:55 + 640] = img
    camUI[44:44 + 633, 808:808 + 414] = imgFilesList[imgFileType]

    if facesFrame:
        for encodeFace, faceLocation in zip(eFrame, facesFrame):
            matches = face_recognition.compare_faces(encodeListSearched, encodeFace)
            distance = face_recognition.face_distance(encodeListSearched, encodeFace)
            # print("matches", matches)
            # print("distance", distance)

            matchIndx = np.argmin(distance)
            # print("matchIndx", matchIndx)

            if matches[matchIndx]:
                    # print("Student found")
                    # print(studentsId[matchIndx])
                    y1, x2, y2, x1 = faceLocation
                    y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
                    bbox = 55 + x1, 162 + y1, x2 - x1, y2 - y1
                    camUI = cvzone.cornerRect(camUI, bbox, rt=0, t=1)

                    id = studentsId[matchIndx]

                    # Define the folder path to save the attendance file
                    attendance_folder = 'attendance_records'

                    # Create the folder if it doesn't exist
                    if not os.path.exists(attendance_folder):
                        os.makedirs(attendance_folder)

                    student_info = db.reference(f'Students/{id}').get()
                    datetime_now = datetime.now()
                    current_date = datetime_now.strftime("%Y-%m-%d")
                    attendance_file_name = f"attendance_record_{current_date}.xlsx"
                    attendance_file_path = os.path.join(attendance_folder, attendance_file_name)

                    # Check if the attendance file exists for today
                    if not os.path.exists(attendance_file_path):
                        # Create a new Excel file for today
                        attendance_workbook = Workbook()
                        attendance_sheet = attendance_workbook.active
                        attendance_sheet.append(["ID", "Name", "Program", "Block&Year", "Total Attendance", "Date/Time"])
                    else:
                        # Load the existing Excel file for today
                        attendance_workbook = openpyxl.load_workbook(attendance_file_path)
                        attendance_sheet = attendance_workbook.active

                        # Check if the student has already attended today
                        existing_attendance_records = attendance_sheet.values
                        existing_attendance_records = [record for record in existing_attendance_records if record[0] == id]

                        if existing_attendance_records:
                            print("Student has already attended today.")
                            attendance_workbook.close()

                    # Record student attendance
                    attendance_sheet.append(
                        [id, student_info['name'], student_info['Program'], student_info['block&year'],
                         student_info['total_attendance'], datetime_now.strftime("%Y-%m-%d %H:%M:%S")])
                    attendance_workbook.save(attendance_file_path)

                    # Update the last attendance time in the dictionary
                    last_attendance_time_dict[id] = datetime_now

                    # imgfiletype start for ui
                    if counter == 0:
                        font_face = cv2.FONT_HERSHEY_SIMPLEX
                        font_scale = 2.5
                        cv2.putText(camUI, "Loading", (100, 400), font_face, font_scale, (255, 0, 0), 2)
                        cv2.waitKey(1)
                        counter = 1
                        imgFileType = 1

            if counter != 0:

                if counter == 1:
                    # data
                    studentsIdInfo = db.reference(f'Students/{id}').get()
                    print(studentsIdInfo)

                    # get image
                    blob = bucket.blob(f'Images/{id}.png')
                    array = np.frombuffer(blob.download_as_string(), np.uint8)
                    imageStudents = cv2.imdecode(array, cv2.COLOR_BGR2RGB)

                    # data of attendance

                    datetimeObject = datetime.strptime(studentsIdInfo['recent_attendance'],
                                                       "%Y-%m-%d %H:%M:%S")
                    secondsElapsed = (datetime.now() - datetimeObject).total_seconds()
                    print(secondsElapsed)
                    if secondsElapsed > 36000:
                        ref = db.reference(f'Students/{id}')
                        studentsIdInfo['total_attendance'] = int(studentsIdInfo.get('total_attendance', 0)) + 1

                        ref.child('total_attendance').set(studentsIdInfo['total_attendance'])
                        ref.child('recent_attendance').set(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                    else:
                        imgFileType = 3
                        counter = 0
                        camUI[44:44 + 633, 808:808 + 414] = imgFilesList[imgFileType]

            if imgFileType != 3:

                if 10 < counter < 20:
                    imgFileType = 2

                camUI[44:44 + 633, 808:808 + 414] = imgFilesList[imgFileType]

                if counter <= 10:
                    cv2.putText(camUI, str(studentsIdInfo['Program']), (1006, 550),
                                cv2.FONT_HERSHEY_COMPLEX, 0.6, (50, 50, 50), 1)
                    cv2.putText(camUI, str(studentsIdInfo['block&year']), (1006, 610),
                                cv2.FONT_HERSHEY_COMPLEX, 0.6, (50, 50, 50), 1)
                    cv2.putText(camUI, str(id), (1006, 493),
                                cv2.FONT_HERSHEY_COMPLEX, 0.6, (50, 50, 50), 1)

                    (w, h), _ = cv2.getTextSize(studentsIdInfo['name'], cv2.FONT_HERSHEY_COMPLEX, 1, 1)
                    offset = (440 - w) // 2
                    cv2.putText(camUI, str(studentsIdInfo['name']), (810 + offset, 445),
                                cv2.FONT_HERSHEY_COMPLEX, 0.8, (100, 100, 100), 1)

                    camUI[175:175 + 216, 909:909 + 216] = imageStudents

                counter += 1

                if counter >= 20:
                    counter = 0
                    imgFileType = 0
                    studentsIdInfo = []
                    imageStudents = []
                    camUI[44:44 + 633, 808:808 + 414] = imgFilesList[imgFileType]
    else:
        imgFileType = 0
        counter = 0

    #cv2.imshow("attendance", img)
    cv2.imshow("Student-face-cam", camUI)
    cv2.waitKey(1)
